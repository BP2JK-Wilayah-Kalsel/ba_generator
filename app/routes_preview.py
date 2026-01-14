import os
import json
import logging
from flask import Blueprint, request, jsonify

logger = logging.getLogger(__name__)
bp = Blueprint('preview', __name__)


@bp.route('/api/preview_document', methods=['POST'])
def preview_document():
    try:
        import mammoth

        data = request.json
        doc_code = data.get('doc_code', '')
        keywords = data.get('keywords', {})
        master_folder = data.get('master_folder', '')

        if not doc_code or not master_folder:
            return jsonify({'success': False, 'error': 'Missing doc_code or master_folder'}), 400

        # Normalize doc_code
        normalized_code = doc_code.replace('_', '-')

        special_files = {
            'DH': ['!Daftar Hadir.docx', 'Daftar Hadir.docx', '!DH.docx'],
            '22-lhp': ['22-Format----------LHP.docx', '22-LHP.docx', '22-Format-LHP.docx'],
            '22_lhp': ['22-Format----------LHP.docx', '22-LHP.docx', '22-Format-LHP.docx']
        }

        try:
            all_files = os.listdir(master_folder)
        except Exception as e:
            logger.exception('Cannot access master folder')
            return jsonify({'success': False, 'error': f'Cannot access master folder: {str(e)}'}), 500

        doc_path = None
        doc_code_lower = doc_code.lower()
        normalized_lower = normalized_code.lower()

        if doc_code_lower in special_files or normalized_lower in special_files:
            special_names = special_files.get(doc_code_lower) or special_files.get(normalized_lower)
            for special_name in special_names:
                if special_name in all_files:
                    doc_path = os.path.join(master_folder, special_name)
                    break

        if not doc_path:
            for filename in all_files:
                if not filename.lower().endswith('.docx'):
                    continue
                if filename.startswith(doc_code) or filename.startswith(normalized_code):
                    doc_path = os.path.join(master_folder, filename)
                    break

        if not doc_path:
            error_msg = f'Dokumen tidak ditemukan untuk code: {doc_code}\nMaster folder: {master_folder}'
            return jsonify({'success': False, 'error': error_msg}), 404

        with open(doc_path, 'rb') as docx_file:
            result = mammoth.convert_to_html(docx_file)
            html_content = result.value
            messages = result.messages

        for key, value in keywords.items():
            placeholder = '{' + key + '}'
            html_content = html_content.replace(placeholder, str(value) if value else '')

        styled_html = f"""
        <style>
            .docx-preview {{ font-family: 'Calibri', 'Arial', sans-serif; font-size: 11pt; line-height: 1.5; color: #000; padding: 20px; background: white; }}
            .docx-preview p {{ margin: 6pt 0; }}
            .docx-preview table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
            .docx-preview table td, .docx-preview table th {{ border: 1px solid #000; padding: 5px; }}
        </style>
        <div class="docx-preview">{html_content}</div>
        """

        remaining_placeholders = list(set([m.group(1) for m in __import__('re').finditer(r'\{([a-zA-Z0-9_\-]+)\}', html_content)])) if html_content else []

        return jsonify({'success': True, 'doc_code': doc_code, 'html': styled_html, 'warnings': [str(m) for m in messages] if messages else [], 'remaining_placeholders': remaining_placeholders, 'doc_path': doc_path})

    except ImportError:
        return jsonify({'success': False, 'error': 'Library mammoth belum terinstall. Jalankan: pip install mammoth'}), 500
    except Exception as e:
        logger.exception('Preview document error')
        return jsonify({'success': False, 'error': f'Terjadi kesalahan: {str(e)}'}), 500


@bp.route('/api/preview_excel', methods=['POST'])
def preview_excel():
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        import tempfile
        import shutil

        data = request.json
        file_path = data.get('file_path', '')
        keywords = data.get('keywords', {})
        company_data = data.get('company_data', {})
        pengalaman_data = data.get('pengalaman_data', {})

        if not file_path or not os.path.exists(file_path):
            return jsonify({'success': False, 'error': f'File Excel tidak ditemukan: {file_path}'}), 404

        temp_dir = tempfile.mkdtemp()
        temp_file = os.path.join(temp_dir, 'preview_temp.xlsx')
        shutil.copy2(file_path, temp_file)

        # Use excel processor from app
        from .excel_processor import fill_excel_pengalaman
        fill_excel_pengalaman(temp_file, company_data, pengalaman_data, {'keywords': keywords})

        wb = load_workbook(temp_file, data_only=True)
        sheet = wb.active
        sheet_name = sheet.title

        html_table = '<table class="table table-bordered table-sm">'
        merged_ranges = list(sheet.merged_cells.ranges)
        processed_cells = set()
        max_row = sheet.max_row
        max_col = sheet.max_column

        for row_idx in range(1, max_row + 1):
            if sheet.row_dimensions[row_idx].hidden:
                continue
            html_table += '<tr>'
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                if sheet.column_dimensions[col_letter].hidden:
                    continue
                cell = sheet.cell(row_idx, col_idx)
                cell_coord = f"{col_letter}{row_idx}"
                if cell_coord in processed_cells:
                    continue
                colspan = 1
                rowspan = 1
                for merged_range in merged_ranges:
                    if cell.coordinate in merged_range:
                        min_col, min_row, max_col_merge, max_row_merge = merged_range.bounds
                        colspan = max_col_merge - min_col + 1
                        rowspan = max_row_merge - min_row + 1
                        for r in range(min_row, max_row_merge + 1):
                            for c in range(min_col, max_col_merge + 1):
                                processed_cells.add(f"{get_column_letter(c)}{r}")
                        break
                cell_value = cell.value if cell.value is not None else ''
                if isinstance(cell_value, str):
                    for key, value in keywords.items():
                        placeholder = '{' + key + '}'
                        cell_value = cell_value.replace(placeholder, str(value) if value else '')
                style_parts = []
                if cell.alignment and cell.alignment.horizontal:
                    style_parts.append(f"text-align: {cell.alignment.horizontal}")
                if cell.alignment and cell.alignment.vertical:
                    style_parts.append(f"vertical-align: {cell.alignment.vertical}")
                if cell.font:
                    if cell.font.bold:
                        style_parts.append("font-weight: bold")
                    if cell.font.italic:
                        style_parts.append("font-style: italic")
                    if cell.font.size:
                        style_parts.append(f"font-size: {cell.font.size}pt")
                try:
                    rgb = cell.fill.start_color.rgb
                    if isinstance(rgb, str) and rgb != '00000000':
                        if len(rgb) == 8:
                            rgb = rgb[2:]
                        style_parts.append(f"background-color: #{rgb}")
                except Exception:
                    pass
                style_parts.append("border: 1px solid #dee2e6")
                style_parts.append("padding: 5px")
                style_attr = f' style="{"; ".join(style_parts)}"' if style_parts else ''
                colspan_attr = f' colspan="{colspan}"' if colspan > 1 else ''
                rowspan_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
                html_table += f'<td{colspan_attr}{rowspan_attr}{style_attr}>{cell_value}</td>'
            html_table += '</tr>'
        html_table += '</table>'

        styled_html = f"""
        <style>
            .excel-preview {{ font-family: 'Calibri', Arial, sans-serif; font-size: 11pt; overflow-x: auto; }}
            .excel-preview table {{ border-collapse: collapse; width: 100%; margin: 0; }}
            .excel-preview td {{ white-space: pre-wrap; word-wrap: break-word; }}
        </style>
        <div class="excel-preview">
            <div class="alert alert-info mb-2"><i class="fas fa-file-excel me-2"></i>Sheet: <strong>{sheet_name}</strong></div>
            {html_table}
        </div>
        """

        remaining_placeholders = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    placeholders = __import__('re').findall(r'\{([a-zA-Z0-9_\-]+)\}', cell.value)
                    remaining_placeholders.extend(placeholders)

        wb.close()
        try:
            import shutil
            shutil.rmtree(temp_dir)
        except Exception:
            pass

        return jsonify({'success': True, 'html': styled_html, 'sheet_name': sheet_name, 'remaining_placeholders': list(set(remaining_placeholders)), 'file_path': file_path})

    except Exception as e:
        logger.exception('Excel preview failed')
        return jsonify({'success': False, 'error': f'Error generating preview: {str(e)}'}), 500
