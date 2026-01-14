import logging
from typing import Dict, Any
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def fill_excel_pengalaman(excel_path: str, company_data: Dict[str, Any], pengalaman_data: Dict[str, Any], form_data: Dict[str, Any]) -> bool:
    try:
        wb = load_workbook(excel_path)

        kso_anggota2 = company_data.get('kso_anggota2', '').strip()
        kso_anggota3 = company_data.get('kso_anggota3', '').strip()

        if not kso_anggota2:
            selected_sheet_name = 'Sheet1'
            sheets_to_delete = ['Sheet2', 'Sheet3']
        elif not kso_anggota3:
            selected_sheet_name = 'Sheet2'
            sheets_to_delete = ['Sheet1', 'Sheet3']
        else:
            selected_sheet_name = 'Sheet3'
            sheets_to_delete = ['Sheet1', 'Sheet2']

        for sheet_name in sheets_to_delete:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        ws = wb[selected_sheet_name]

        sejenis_count = int(pengalaman_data.get('sejenis', 7))
        tahun_sejenis = int(pengalaman_data.get('tahun_sejenis', 10))
        beda_jenis_count = int(pengalaman_data.get('beda_jenis', 4))
        tahun_beda_jenis = int(pengalaman_data.get('tahun_beda_jenis', 4))

        SEJENIS_START_ROW = 7
        SEJENIS_MAX_ROWS = 15
        BEDA_JENIS_START_ROW = 23
        BEDA_JENIS_MAX_ROWS = 10

        for r in range(SEJENIS_START_ROW, SEJENIS_START_ROW + SEJENIS_MAX_ROWS):
            ws.row_dimensions[r].hidden = False
        if sejenis_count < 0:
            sejenis_count = 0
        if sejenis_count > SEJENIS_MAX_ROWS:
            sejenis_count = SEJENIS_MAX_ROWS
        for r in range(SEJENIS_START_ROW + sejenis_count, SEJENIS_START_ROW + SEJENIS_MAX_ROWS):
            ws.row_dimensions[r].hidden = True

        for r in range(BEDA_JENIS_START_ROW, BEDA_JENIS_START_ROW + BEDA_JENIS_MAX_ROWS):
            ws.row_dimensions[r].hidden = False
        if beda_jenis_count < 0:
            beda_jenis_count = 0
        if beda_jenis_count > BEDA_JENIS_MAX_ROWS:
            beda_jenis_count = BEDA_JENIS_MAX_ROWS
        for r in range(BEDA_JENIS_START_ROW + beda_jenis_count, BEDA_JENIS_START_ROW + BEDA_JENIS_MAX_ROWS):
            ws.row_dimensions[r].hidden = True

        leadfirm = company_data.get('leadfirm', '')
        nama_kso = company_data.get('nama_kso', '')
        note_pengalaman = (
            (form_data.get('note_pengalaman') if isinstance(form_data, dict) else None)
            or (form_data.get('keywords', {}).get('note_pengalaman') if isinstance(form_data, dict) else None)
            or ''
        )

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell.value = (cell.value
                        .replace('{X_tahun_sejenis}', str(tahun_sejenis))
                        .replace('{X_tahun_beda_jenis}', str(tahun_beda_jenis))
                        .replace('{note_pengalaman}', note_pengalaman)
                        .replace('{leadfirm}', leadfirm)
                        .replace('{nama_kso}', nama_kso)
                        .replace('{kso_anggota2}', kso_anggota2)
                        .replace('{kso_anggota3}', kso_anggota3)
                    )

        wb.save(excel_path)
        wb.close()
        return True

    except Exception as e:
        logger.exception(f"Error filling Excel: {e}")
        return False
