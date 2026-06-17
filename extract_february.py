
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def extract_february_data():
    file_path = r'D:\2025\SC\ba_generator\static\template_response\data tabel.html'
    output_path = r'D:\2025\SC\ba_generator\data_februari.xlsx'
    
    if not os.path.exists(file_path):
        print(f"Error: File {file_path} tidak ditemukan.")
        return

    print(f"Membaca file {file_path}...")
    with open(file_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')

    # Find the table headers
    headers = []
    thead = soup.find('thead')
    if thead:
        header_cells = thead.find_all('th')
        headers = [cell.get_text(strip=True) for cell in header_cells]
    
    # Identify column index for "Tanggal Penetapan"
    # Based on observation, it should be index 11
    target_col_name = "Tanggal Penetapan"
    try:
        target_idx = headers.index(target_col_name)
    except ValueError:
        # Fallback to index 11 if header name match fails
        target_idx = 11
        print(f"Warning: Kolom '{target_col_name}' tidak ditemukan di header. Menggunakan index default {target_idx}.")
    else:
        print(f"Kolom '{target_col_name}' ditemukan pada index {target_idx}.")

    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Februari"

    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Write headers to Excel
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Filter rows from tbody
    tbody = soup.find('tbody')
    rows_extracted = 0
    if tbody:
        rows = tbody.find_all('tr')
        print(f"Total baris dalam tabel: {len(rows)}")
        
        current_excel_row = 2
        for tr in rows:
            tds = tr.find_all('td')
            if len(tds) > target_idx:
                tanggal_penetapan = tds[target_idx].get_text(strip=True)
                
                # Check if it contains "Februari"
                if "Februari" in tanggal_penetapan:
                    # Collect all data from this row
                    for col_num, td in enumerate(tds, 1):
                        # For Nama Paket (index 2) and others, handle nested text
                        cell_value = td.get_text(separator=" ", strip=True)
                        cell = ws.cell(row=current_excel_row, column=col_num, value=cell_value)
                        cell.border = thin_border
                    
                    current_excel_row += 1
                    rows_extracted += 1

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

    # Save workbook
    wb.save(output_path)
    print(f"Berhasil mengekstrak {rows_extracted} baris data ke {output_path}")

if __name__ == "__main__":
    extract_february_data()
